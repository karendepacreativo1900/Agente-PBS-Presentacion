#!/usr/bin/env python3
"""Script para generar un Excel de ejemplo con datos de ventas."""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()

# Hoja 1: Resumen Ejecutivo
ws1 = wb.active
ws1.title = "Resumen Ejecutivo"
ws1["A1"] = "Indicador"
ws1["B1"] = "2023"
ws1["C1"] = "2024"
ws1["D1"] = "Variación %"
datos_resumen = [
    ("Ingresos Totales (M$)", 4.2, 5.8, 38),
    ("Clientes Activos", 1200, 1850, 54),
    ("Satisfacción Cliente (%)", 82, 91, 11),
    ("Nuevos Productos", 3, 7, 133),
    ("Empleados", 45, 62, 38),
]
for row in datos_resumen:
    ws1.append(row)

# Hoja 2: Ventas por Región
ws2 = wb.create_sheet("Ventas por Región")
ws2.append(["Región", "Q1", "Q2", "Q3", "Q4", "Total Anual"])
ventas = [
    ("Norte", 320, 410, 390, 480, 1600),
    ("Sur", 210, 280, 310, 350, 1150),
    ("Centro", 450, 520, 490, 610, 2070),
    ("Este", 180, 220, 260, 300, 960),
    ("Oeste", 290, 340, 380, 420, 1430),
]
for row in ventas:
    ws2.append(row)

# Hoja 3: Productos Top
ws3 = wb.create_sheet("Productos Top")
ws3.append(["Producto", "Unidades Vendidas", "Ingresos ($)", "Margen (%)", "Calificación"])
productos = [
    ("Producto Alpha", 4500, 900000, 42, 4.8),
    ("Producto Beta", 3200, 640000, 38, 4.5),
    ("Producto Gamma", 2800, 560000, 45, 4.7),
    ("Producto Delta", 1900, 380000, 35, 4.2),
    ("Producto Epsilon", 1500, 300000, 50, 4.9),
]
for row in productos:
    ws3.append(row)

# Hoja 4: Objetivos 2025
ws4 = wb.create_sheet("Objetivos 2025")
ws4.append(["Objetivo", "Meta", "Plazo", "Responsable"])
objetivos = [
    ("Incrementar ingresos", "30%", "Dic 2025", "Dirección Comercial"),
    ("Expansión nuevas regiones", "3 regiones", "Jun 2025", "Expansión"),
    ("Lanzar nuevos productos", "5 productos", "Sep 2025", "I+D"),
    ("Reducir tiempos de entrega", "20%", "Mar 2025", "Operaciones"),
    ("Mejorar NPS", "Score 70+", "Dic 2025", "Experiencia Cliente"),
]
for row in objetivos:
    ws4.append(row)

out_path = Path(__file__).parent / "datos_ejemplo.xlsx"
wb.save(out_path)
print(f"Excel de ejemplo creado: {out_path}")
